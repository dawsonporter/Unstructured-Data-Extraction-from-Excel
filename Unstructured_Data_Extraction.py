import os
import openpyxl
import tkinter as tk
from tkinter import filedialog, ttk, StringVar, Toplevel
from openpyxl import Workbook
from collections import defaultdict
from tkinter import StringVar
import difflib
import csv
import re
from collections import deque
from threading import Timer
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
import nltk
nltk.download('punkt')
nltk.download('stopwords')

# Create GUI interface
root = tk.Tk()
root.configure(bg='#ADD8E6')  # 'Light Blue' color
root.geometry('1800x1000')  # Adjusted the GUI size

canvas = tk.Canvas(root, bg='#ADD8E6')
canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

scrollbar = tk.Scrollbar(root, command=canvas.yview)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

scrollable_frame = tk.Frame(canvas, bg='#ADD8E6') 
scrollable_frame.bind(
    "<Configure>",
    lambda e: canvas.configure(
        scrollregion=canvas.bbox("all")
    )
)

canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
canvas.configure(yscrollcommand=scrollbar.set)

# Initialize variables
frames = []
search_texts = []
num_values = []
directions = []
sheets = []
directory = StringVar()
sheet_options = []
column_names = []
keywords = []

timers = {}

def format_value(value):
    return str(value)

def process_files():
    summary_filename = "0 - summary.csv"

    with open(os.path.join(directory.get(), summary_filename), 'w', newline='') as summary_file:
        writer = csv.writer(summary_file)

        # First determine the max occurrences for each keyword
        max_occurrences = defaultdict(int)

        for filename in os.listdir(directory.get()):
            if filename.startswith('~$') or filename == summary_filename:
                continue
            if filename.endswith(".xlsx") or filename.endswith(".xls"):
                filepath = os.path.join(directory.get(), filename)
                workbook = openpyxl.load_workbook(filepath, data_only=True)
                right_values = defaultdict(list)
                
                target_sheets = [sheet.get() for sheet in sheets]

                for target_sheet in target_sheets:
                    if target_sheet in [sheet.title for sheet in workbook.worksheets]:
                        process_sheet(workbook[target_sheet], right_values)

                for search_text, num_value, column_name in zip(search_texts, num_values, column_names):
                    unique_key = f"{search_text.get()}_{num_value.get()}"  # make a unique key
                    max_occurrences[unique_key] = max(max_occurrences[unique_key], len(right_values[unique_key]))

        # Create the headers
        headers = ['Filename']
        for search_text, num_value, column_name in zip(search_texts, num_values, column_names):
            unique_key = f"{search_text.get()}_{num_value.get()}"  # make a unique key
            column_label = column_name.get() if column_name.get() != '' else search_text.get()
            headers.extend([f"{column_label}{i+1}" for i in range(max_occurrences[unique_key])])

        writer.writerow(headers)

        # Write the rows
        for filename in os.listdir(directory.get()):
            if filename.startswith('~$') or filename == summary_filename:
                continue
            if filename.endswith(".xlsx") or filename.endswith(".xls"):
                filepath = os.path.join(directory.get(), filename)
                workbook = openpyxl.load_workbook(filepath, data_only=True)
                right_values = defaultdict(list)

                target_sheets = [sheet.get() for sheet in sheets]
                for target_sheet in target_sheets:
                    if target_sheet in [sheet.title for sheet in workbook.worksheets]:
                        process_sheet(workbook[target_sheet], right_values)

                row = [filename]
                for search_text, num_value, column_name in zip(search_texts, num_values, column_names):
                    unique_key = f"{search_text.get()}_{num_value.get()}"  # make a unique key

                    # append values and then fill remaining with None
                    row.extend(right_values[unique_key])
                    row.extend([None] * (max_occurrences[unique_key] - len(right_values[unique_key])))

                writer.writerow(row)

    extraction_complete_label = tk.Label(top_frame, text="Data extraction complete", fg='black', bg='#808080')
    extraction_complete_label.pack()
    root.after(2000, extraction_complete_label.pack_forget)  # Remove the label after 2 seconds

def process_sheet(sheet, right_values):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                for search_text, num_value, direction in zip(search_texts, num_values, directions):
                    unique_key = f"{search_text.get()}_{num_value.get()}"  # make a unique key
                    if re.search(search_text.get().lower(), str(cell.value).lower()):
                        n_values = int(num_value.get())
                        values_found = 0
                        offset = 1 if direction.get() in ('Right', 'Down') else -1
                        while cell.column + offset <= sheet.max_column and values_found < n_values:
                            next_cell = sheet.cell(row=cell.row + offset if direction.get() in ('Up', 'Down') else cell.row, column=cell.column + offset if direction.get() in ('Right', 'Left') else cell.column)
                            if next_cell.value is not None and str(next_cell.value).strip() != '':
                                values_found += 1
                                if values_found == n_values:
                                    keywords = str(next_cell.value).split(",")  # Split the keywords
                                    sorted_keywords = sorted([keyword.strip() for keyword in keywords])  # Sort and remove leading/trailing spaces
                                    right_values[unique_key].extend(sorted_keywords)  # Add the sorted keywords
                            offset += 1 if direction.get() in ('Right', 'Down') else -1

def get_keywords():
    stop_words = set(stopwords.words('english')) 
    unique_words = set()
    combined_keywords = set()
    
    target_sheets = [sheet.get() for sheet in sheets]

    for filename in os.listdir(directory.get()):
        if filename.startswith('~$'):
            continue
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            filepath = os.path.join(directory.get(), filename)
            workbook = openpyxl.load_workbook(filepath, data_only=True)

            for sheet in workbook.worksheets:
                if sheet.title not in target_sheets:
                    continue
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None and isinstance(cell.value, str):
                            words = word_tokenize(cell.value)
                            meaningful_words = [word for word in words if word.isalnum() and word not in stop_words] 
                            for word in meaningful_words:
                                unique_words.add(word)
                            if len(meaningful_words) > 1:
                                combined_keyword = "_".join(meaningful_words)
                                combined_keywords.add(combined_keyword)
    
    return sorted(list(unique_words)), sorted(list(combined_keywords))
                                                      
def add_search_text(keyword=None, direction=None, values=None, keywords=None):
    search_text = StringVar()
    number_of_values = StringVar()
    direction_str = StringVar()
    column_name = StringVar()

    if keyword and direction and values:
        search_text.set(keyword)
        number_of_values.set(str(values))  # Convert values to string before setting
        direction_str.set(direction)
        column_name.set(keyword) # Set the default column name as the keyword
    else:
        search_text.set("")
        number_of_values.set("1")
        direction_str.set("Right")
        column_name.set("")

    search_texts.append(search_text)
    num_values.append(number_of_values)
    directions.append(direction_str)
    column_names.append(column_name) # append the new column name variable to the list

    frame = tk.Frame(scrollable_frame, bg='#808080')  # 'Gray' color
    frame.pack(pady=10, fill=tk.X)
    frames.append(frame)

    label1 = tk.Label(frame, text='Keyword', fg='black', bg='#808080')
    label1.grid(row=0, column=0)

    label2 = tk.Label(frame, text='Where is Value?', fg='black', bg='#808080')
    label2.grid(row=0, column=1)

    label3 = tk.Label(frame, text='# of Values Away?', fg='black', bg='#808080')
    label3.grid(row=0, column=2)

    # Add new label for column names
    label4 = tk.Label(frame, text='Name Column (Optional)', fg='black', bg='#808080')
    label4.grid(row=0, column=3)

    # Use a Combobox instead of an Entry for the keyword input
    search_text_entry = ttk.Combobox(frame, textvariable=search_text, width=70)
    search_text_entry['values'] = keywords
    search_text_entry.grid(row=1, column=0)
    search_text_entry.bind('<KeyRelease>', lambda e: update_combobox(search_text_entry, keywords)) # bind <KeyRelease> event to update_combobox function

    direction_entry = ttk.Combobox(frame, textvariable=direction_str, values=['Up', 'Down', 'Right', 'Left'], state='readonly')
    direction_entry.grid(row=1, column=1)

    number_of_values_entry = ttk.Combobox(frame, textvariable=number_of_values, values=[str(i) for i in range(1, 11)], state='readonly')
    number_of_values_entry.grid(row=1, column=2)

    # Add new entry for column names
    column_name_entry = ttk.Combobox(frame, textvariable=column_name, width=30)
    column_name_entry['values'] = keywords
    column_name_entry.grid(row=1, column=3)
    column_name_entry.bind('<KeyRelease>', lambda e: update_combobox(column_name_entry, keywords)) # bind <KeyRelease> event to update_combobox function

    delete_button = tk.Button(frame, text="Delete", fg='black', bg='#808080', 
                              command=lambda: delete_search_text(frame, search_text, number_of_values, direction_str, column_name))  # Delete button added
    delete_button.grid(row=1, column=4)

def update_combobox(combobox, keywords):
    global timers

    # if a call is already scheduled for this combobox, cancel it
    if combobox in timers:
        timers[combobox].cancel()

    # schedule the actual_update_combobox call
    timers[combobox] = Timer(0.5, actual_update_combobox, [combobox, keywords])
    timers[combobox].start()

def actual_update_combobox(combobox, keywords):
    # current text in combobox
    value = combobox.get()
    value = value.lower()

    # use list comprehension to find matching items
    matching_items = [item for item in keywords if item.lower().startswith(value)]

    # update combobox values
    combobox['values'] = matching_items

    if matching_items:
        # avoid popping up dropdown if input field is empty
        if value != '':
            combobox.event_generate('<Down>')  # shows the dropdown menu
    else:
        combobox.event_generate('<Escape>')  # hides the dropdown menu if no matching items

    # keep the focus on the combobox
    combobox.focus_set()  # sets focus back to the combobox

    # restore the original input after dropdown update
    combobox.set(value)
    combobox.icursor(tk.END)  # move cursor to the end of input
  
    
def delete_search_text(frame, search_text, number_of_values, direction_str, column_name):  # Function to delete text fields
    frames.remove(frame)
    search_texts.remove(search_text)
    num_values.remove(number_of_values)
    directions.remove(direction_str)
    column_names.remove(column_name)
    frame.destroy()

def browse_files():
    filename = filedialog.askdirectory()
    directory.set(filename)
    dir_label = tk.Label(top_frame, text=f"Selected Directory: {filename}", fg='black', bg='#ADD8E6')
    dir_label.pack()

def load_sheets():
    global sheets  # ensure we are modifying the global variable
    global sheet_options  # ensure we are modifying the global variable
    
    # Clear the existing sheets
    sheets.clear()
    sheet_options.clear()

    # Iterate over each file in the directory
    for filename in os.listdir(directory.get()):
        if filename.startswith('~$'):
            continue
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            filepath = os.path.join(directory.get(), filename)
            workbook = openpyxl.load_workbook(filepath, data_only=True)
            
            # Add sheet names of the workbook to sheets list
            for sheet in workbook.sheetnames:
                if sheet not in sheet_options:
                    sheet_options.append(sheet)

    # Sort sheet_options in alphabetical order
    sheet_options.sort()

    load_sheets_label = tk.Label(top_frame, text="Sheets loaded complete", fg='black', bg='#ADD8E6')
    load_sheets_label.pack()

def load_keywords():
    global keywords
    keywords.clear()
    combined_keywords = set()  # Set to store combined keywords
    selected_sheets = [sheet.get() for sheet in sheets]  # Get the sheet names from the StringVar objects

    # Iterate over each file in the directory
    for filename in os.listdir(directory.get()):
        if filename.startswith('~$'):
            continue
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            filepath = os.path.join(directory.get(), filename)
            workbook = openpyxl.load_workbook(filepath, data_only=True)

            # Iterate over each selected sheet in that file
            for sheet in selected_sheets:
                if sheet in workbook.sheetnames:
                    worksheet = workbook[sheet]

                    # Extract the unique keywords from these selected sheets
                    for row in worksheet.iter_rows(values_only=True):
                        for value in row:
                            if value is not None and isinstance(value, str):
                                # Tokenize the cell value
                                words = word_tokenize(value)
                                # Add each word as a keyword if it does not contain a number or special character
                                for word in words:
                                    if word.isalnum():
                                        keywords.append(word)

                                # Add the whole cell value (after lowercasing and removing leading/trailing spaces) as a combined keyword
                                combined_keywords.add(value.lower().strip())

    # Convert keywords to a set and back to list to remove duplicates
    keywords = list(set(keywords))

    # Sort the single keywords
    keywords.sort()

    # Convert combined keywords to list and sort them
    combined_keywords = sorted(list(combined_keywords))

    # Add combined keywords to the keywords list
    keywords.extend(combined_keywords)

    load_keywords_label = tk.Label(top_frame, text="Keywords loaded successfully", fg='black', bg='#808080')
    load_keywords_label.pack()
    root.after(2000, load_keywords_label.pack_forget)  # Remove the label after 2 seconds
    
def add_sheet_name(sheet_name=None):
    sheet = StringVar()

    if sheet_name:
        sheet.set(sheet_name)
    else:
        sheet.set("")

    sheets.append(sheet)

    frame = tk.Frame(scrollable_frame, bg='#808080')  # 'Gray' color
    frame.pack(pady=10, fill=tk.X)
    frames.append(frame)

    sheet_name_entry = ttk.Combobox(frame, textvariable=sheet, values=sheet_options, state='readonly')  # Background color white
    sheet_name_entry.grid(row=1, column=0)
    sheet_name_entry.bind('<KeyRelease>', lambda e: update_combobox(sheet_name_entry, sheet_options))

    delete_button = tk.Button(frame, text="Delete", fg='black', bg='#808080', command=lambda: delete_sheet_name(frame, sheet))  # Delete button added
    delete_button.grid(row=1, column=1)

def delete_sheet_name(frame, sheet_name):  # Function to delete sheet fields
    frames.remove(frame)
    sheets.remove(sheet_name)
    frame.destroy()

top_frame = tk.Frame(root, bg='#ADD8E6')  
top_frame.pack(fill=tk.X)

step_labels = ["Step 1: Select Folder",
               "Step 2: Load Sheets",
               "Step 3: Add Sheet",
               "Step 4: Load Keywords",
               "Step 5: Add Keyword",
               "Step 6: Extract Data / Run Summary"]

step_descriptions = ["Choose the folder where your files are located",
                     "Load unique sheets from the Excel files",
                     "Add sheet(s) you want to extract data from",
                     "Load keywords from the selected sheets",
                     "Add keyword & details you want to search for in the sheet",
                     "Start the data extraction and create a csv summary in the file setup as directory"]

buttons = [tk.Button(top_frame, text="Select Directory", command=browse_files, bg='yellow', fg='black'),
           tk.Button(top_frame, text="Load Sheets", command=load_sheets, bg='yellow', fg='black'),
           tk.Button(top_frame, text="Add Sheet", command=add_sheet_name, bg='yellow', fg='black'),
           tk.Button(top_frame, text="Load Keywords", command=load_keywords, bg='yellow', fg='black'),
           tk.Button(top_frame, text="Add Keyword", command=lambda: add_search_text(keywords=keywords), bg='yellow', fg='black'),
           tk.Button(top_frame, text="Extract Data / Run Summary", command=process_files, bg='yellow', fg='black')]

for i in range(len(buttons)):
    tk.Label(top_frame, text=step_labels[i], fg='black', bg='#ADD8E6', font=('Helvetica', 12, 'bold')).pack(pady=10)
    tk.Label(top_frame, text=step_descriptions[i], fg='black', bg='#ADD8E6', font=('Helvetica', 10, 'italic')).pack()
    buttons[i].pack(pady=10)
    tk.Frame(top_frame, height=2, bg="black").pack(fill=tk.X, pady=10)

root.mainloop()