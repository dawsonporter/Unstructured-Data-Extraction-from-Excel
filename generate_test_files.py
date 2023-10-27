import random
from openpyxl import Workbook
import os

# Setup directory on desktop to save the files
desktop = os.path.expanduser("~") + "/Desktop/"
folder_name = "TestExcelFiles"
folder_path = desktop + folder_name

# Create directory if not exists
if not os.path.exists(folder_path):
    os.mkdir(folder_path)

# Random data generators
def random_name():
    first_names = ["John", "Jane", "Alice", "Bob", "Charlie", "David", "Eve", "Frank"]
    last_names = ["Smith", "Johnson", "Brown", "Williams", "Jones", "Miller", "Davis", "Garcia"]
    return random.choice(first_names) + " " + random.choice(last_names)

def random_age():
    return random.randint(20, 65)

def random_address():
    streets = ["Main St", "High St", "Elm St", "Oak St", "Pine St"]
    cities = ["New York", "Los Angeles", "Chicago", "Houston", "Phoenix"]
    return f"{random.randint(1,9999)} {random.choice(streets)}, {random.choice(cities)}"

def random_email(name):
    domains = ["example.com", "sample.org", "demo.net", "testmail.com"]
    return name.lower().replace(" ", ".") + "@" + random.choice(domains)

def random_phone():
    return f"{random.randint(100,999)}-{random.randint(100,999)}-{random.randint(1000,9999)}"

# Define the unstructured locations for each key
personal_layout = {
    "Name": (3, 2),
    "Age": (6, 5),
    "Address": (8, 4)
}

contact_layout = {
    "Email": (4, 3),
    "Work Phone": (7, 6),
    "Home Phone": (2, 5)
}

# Generate Excel files with data
for i in range(50):
    # Create a new workbook
    wb = Workbook()

    # Generate personal info for the individual
    name = random_name()
    age = random_age()
    address = random_address()

    # Personal Info sheet
    ws1 = wb.active
    ws1.title = "Personal Info"
    ws1.cell(row=personal_layout["Name"][0], column=personal_layout["Name"][1], value="Name")
    ws1.cell(row=personal_layout["Name"][0], column=personal_layout["Name"][1]+1, value=name)
    ws1.cell(row=personal_layout["Age"][0], column=personal_layout["Age"][1], value="Age")
    ws1.cell(row=personal_layout["Age"][0], column=personal_layout["Age"][1]+1, value=age)
    ws1.cell(row=personal_layout["Address"][0], column=personal_layout["Address"][1], value="Address")
    ws1.cell(row=personal_layout["Address"][0], column=personal_layout["Address"][1]+1, value=address)

    # Generate contact info for the same individual
    email = random_email(name)
    work_phone = random_phone()
    home_phone = random_phone()

    # Contact Info sheet
    ws2 = wb.create_sheet(title="Contact Info")
    ws2.cell(row=contact_layout["Email"][0], column=contact_layout["Email"][1], value="Email")
    ws2.cell(row=contact_layout["Email"][0], column=contact_layout["Email"][1]+1, value=email)
    ws2.cell(row=contact_layout["Work Phone"][0], column=contact_layout["Work Phone"][1], value="Work Phone")
    ws2.cell(row=contact_layout["Work Phone"][0], column=contact_layout["Work Phone"][1]+1, value=work_phone)
    ws2.cell(row=contact_layout["Home Phone"][0], column=contact_layout["Home Phone"][1], value="Home Phone")
    ws2.cell(row=contact_layout["Home Phone"][0], column=contact_layout["Home Phone"][1]+1, value=home_phone)

    # Save the workbook
    file_name = f"{name.replace(' ', '_')}_info.xlsx"
    file_path = os.path.join(folder_path, file_name)
    wb.save(file_path)

print(f"Generated 50 test Excel files in {folder_path}/")
